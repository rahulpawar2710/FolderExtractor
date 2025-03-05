from flask import Flask, request, jsonify
from openai import OpenAI
from pathlib import Path
import pandas as pd
import json
import re
import time
import os
from openpyxl import load_workbook
from datetime import datetime
from lxml import etree
import streamlit as st


app = Flask(__name__)


# Function to write or append data to Excel
def write_or_append_excel(df, logFilePath, sheet_name="Sheet1"):
    if not os.path.exists(logFilePath):
        df.to_excel(logFilePath, index=False, engine="openpyxl", sheet_name=sheet_name)
        print(f"File '{logFilePath}' created and data written.")
    else:
        with pd.ExcelWriter(logFilePath, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)
        print(f"Data appended to existing file '{logFilePath}'.")

def get_vector_store_id_by_name(target_name,client):
    try:
        vector_stores = client.beta.vector_stores.list()
        for vector_store in vector_stores:
            if vector_store.name == target_name:
                return vector_store.id
        return None
    except Exception as e:
        print("Error retrieving vector store ID:", e)
        return None

# Function to load XML content
def load_xml_file(file_path):
    with open(file_path, 'rb') as file:
        xml_content = file.read()
    return xml_content

# Function to parse XML content
def parse_xml(xml_content):
    return etree.XML(xml_content)

# Function to get element by full XPath
def get_element_by_full_xpath(tree, xpath_expr):
    elements = tree.xpath(xpath_expr)
    return elements[0] if elements else None
    

def delete_all_files_from_main_storage(client):
  try:
      files = client.files.list()
      # Loop through and delete each file
      for file in files.data:  # Access the list of files using `.data`
                    file_id = file.id
                    print(f"Deleting file: {file_id}")
                    client.files.delete(file_id)
                    print(f"File {file_id} deleted.")
                #response = client.files.delete(file_citation.file_id)
      print("File deleted from Main Storage : ",file_id)
      finalStatus="File deleted from Main Storage"
      return finalStatus
  except Exception as e:
       finalStatus="Issue in file deleteion from Main storage",e
       return finalStatus


def delete_file_from_vector_store(client,vector_store_id,file_id):
    try:
        # Delete the file using OpenAI's API
        client.beta.vector_stores.files.delete(vector_store_id=vector_store_id, file_id=file_id)
        print(f"File with ID {file_id} is being deleted.")
    except OpenAI.error.OpenAIError as e:
        print(f"Error deleting file: {e}")

#@app.route('/process_invoices', methods=['GET'])
def process_invoices(folder_path):
    try:

         # Define necessary variables 
        file_path = 'Config.xml'  # Path to the XML file
        xml_content = load_xml_file(file_path)
        tree = parse_xml(xml_content)
        
        # Reading active environment details from XML
        ActiveEnvironmentPath = '/Configuration/ActiveEnvironment'
        ActiveEnvironmentValue = get_element_by_full_xpath(tree, ActiveEnvironmentPath).text
        os.environ['OPENAI_API_KEY'] = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/OpenAiApiKey").text
        ContextualOpenAiModelName = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/ContextualOpenAiModelName").text
        VectorName = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/VectorName").text
        InputFolderPath = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/InputFolderPath").text
        OutputFolderPath = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/OutputFolderPath").text
        LogFolderPath = get_element_by_full_xpath(tree, f"/Configuration/{ActiveEnvironmentValue}/LogPath").text

        client = OpenAI(api_key=os.environ['OPENAI_API_KEY'])
        assistant = client.beta.assistants.create(
            name="Invoice Processing Assistant",
            instructions="You are an expert Invoice analyst. Use your knowledge base to answer questions.",
            model=ContextualOpenAiModelName,
            tools=[{"type": "file_search"}],
        )

        #Input_folder_path = Path(InputFolderPath)
        Input_folder_path = Path(folder_path)
        vector_store_id = get_vector_store_id_by_name(VectorName,client)
        if not vector_store_id:
            vector_store = client.beta.vector_stores.create(name=VectorName)
            vector_store_id = vector_store.id

        now = datetime.now()
        logFilePath = os.path.join(LogFolderPath, f"Logs_{now.strftime('%d_%b_%Y_%H_%M_%S')}.xlsx")
        OutputFilePath = os.path.join(OutputFolderPath, f"Output_{now.strftime('%d_%b_%Y_%H_%M_%S')}.xlsx")

        
        for file in Input_folder_path.glob("*.pdf"):
            starttime = now.strftime("%Y-%m-%d %H:%M:%S")
            with open(file, "rb") as f:
                file_batch = client.beta.vector_stores.file_batches.upload_and_poll(
                    vector_store_id=vector_store_id, files=[f]
                )
                print(f"File {file} uploaded.")

                time.sleep(7)
                finalStatus = f"File uploaded to Vector Store (id: {vector_store_id})"

                file_id = ""
                vector_store_files = client.beta.vector_stores.files.list(vector_store_id=vector_store_id)

                if not vector_store_files.data:
                    print("No files found in the vector store.")

                latest_file = vector_store_files.data[-1]  # Get the most recent file
                file_id = latest_file.id
                print("Latest File Id : ",file_id)

            assistant = client.beta.assistants.update(
                assistant_id=assistant.id,
                tool_resources={"file_search": {"vector_store_ids": [vector_store_id]}},
            )

            thread = client.beta.threads.create(
                messages=[
                    {
                        "role": "user",
                        "content": (
                            "Please answer the questions in JSON format: "
                            "File Name, Invoice Number, Invoice Date, Purchase Order, Subtotal, HST, Total.\n\n"
                            "I am an invoice processing agent responsible for reading invoices in PDF or image format "
                            "and extracting key fields accurately."
                        )
                    }
                ]
            )

            run = client.beta.threads.runs.create_and_poll(
                thread_id=thread.id, assistant_id=assistant.id
            )

            messages = list(client.beta.threads.messages.list(thread_id=thread.id, run_id=run.id))
            message_content = messages[0].content[0].text
            print(f"Message content: {message_content}")
            
            annotations = message_content.annotations
            citations = []
            for index, annotation in enumerate(annotations):
                message_content.value = message_content.value.replace(annotation.text, f"[{index}]")
                if file_citation := getattr(annotation, "file_citation", None):
                    #cited_file = client.files.retrieve(file_citation.file_id)
                    cited_file=client.files.retrieve(file_citation.file_id)
                    print("Testing File Id : ", file_citation.file_id)
                    citations.append(f"[{index}] {cited_file.filename}")

            result = message_content.value

            match = re.search(r"\{.*\}", result, re.DOTALL)
            if match:
                json_data = match.group(0)
                try:
                    parsed_result = json.loads(json_data)
                    df = pd.json_normalize(parsed_result, sep="_")
                    flat_json = df.to_dict(orient="records")[0]
                    df = pd.DataFrame([flat_json])

                    try:
                        with pd.ExcelWriter(OutputFilePath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                            df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                            finalStatus = "Data extracted and inserted to main file"
                    except FileNotFoundError:
                        df.to_excel(OutputFilePath, index=False, engine="openpyxl")
                        finalStatus = "Data extracted and inserted to main file"

                    response = client.beta.vector_stores.files.delete(vector_store_id=vector_store_id, file_id=file_id)
                    print(f"File deleted from Vector store: {vector_store_id}")
                    finalStatus = "Data extracted, inserted to main file, and file deleted from Vector store"

                    # Delete all files from Main Storage
                    finalStatus = delete_all_files_from_main_storage(client)

                except json.JSONDecodeError as e:
                    print(f"Error parsing JSON: {e}")
                    finalStatus = f"Issues with data extraction, Exception Details: {e}"
                    continue

            time.sleep(5)
            deleted_thread = client.beta.threads.delete(thread.id)

            if hasattr(deleted_thread, 'id'):
                print(f"Thread {deleted_thread.id} has been successfully deleted.")
            else:
                print(f"Failed to delete thread {thread.id}.")

            endTime = now.strftime("%Y-%m-%d %H:%M:%S")
            data1 = {
                "FileName": [file],
                "StartTime": [starttime],
                "EndTime": [endTime],
                "Status": [finalStatus],
                "Comment": [""]
            }

            df = pd.DataFrame(data1)
            write_or_append_excel(df, logFilePath)

        #return jsonify({"status": "success", "message": "Invoices processed successfully"}), 200

        if not os.path.isdir(folder_path):
            return "Invalid folder path!", None, pd.DataFrame()
    
    # Sample extraction logic (Modify as needed)
        files = os.listdir(folder_path)
        output_file = os.path.join(folder_path, "output.xlsx")
    
    # Create a sample DataFrame with SrNo, File Name, and Status
        df = pd.DataFrame({"SrNo": range(1, len(files) + 1), "File Name": files})
    
    # Simulating additional columns (modify as per real data processing)
        df["Status"] = ["Success" if "not found" not in file.lower() else "Fail" for file in files]
    
        df.to_excel(output_file, index=False)
    
        return f"Found {len(files)} files in the folder.", output_file, df

    except Exception as e:
        print(f"Error processing invoices: {e}")
        #return jsonify({"status": "error", "message": str(e)}), 500
        return f"Error processing invoices: {e}"


def main():
    st.title("Folder Data Extractor")
    
    folder_path = st.text_input("Enter the folder path:")
    output_path = ""
    df_result = pd.DataFrame()
    
    if st.button("Extract Data", key="extract_data_btn"):
        if folder_path and os.path.isdir(folder_path):
            result, output_path, df_result = process_invoices(folder_path)
            st.success(result)
        else:
            st.error("Please enter a valid folder path.")
    
    if output_path:
        st.write(f"Output Excel File: {output_path}")
        st.dataframe(df_result)



if __name__ == '__main__':
    #app.run(debug=True)
    main()
